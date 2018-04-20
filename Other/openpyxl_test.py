from openpyxl import workbook, load_workbook
import logging
import logging_conf
LOGGER = logging.getLogger('Trail')


class openpyxl:

    @staticmethod
    def pyxl():
        # print("openpyxl testin")
        wb = load_workbook(filename='/home/stride/PycharmProjects/TrailThings/TestDataX.xlsx')
        # ws = wb.get_sheet_by_name('TestDataS')
        ws = wb['TestDataS']

        # print(ws['A1'].values)

        # print(ws.cell(row=1, column=2).value)

        # ws.cell(row=1, column=2).value = 'Vinodkumar'

        # print(ws.cell(row=1, column=2).value)

        # print(ws.max_row)

        # print(ws.max_column)

        # ws = wb.active
        # column_len = list(ws.columns)[0]
        # print(list(ws.columns)[0].values)
        # print(len(column_len))

        # for obj in list(ws.columns)[1]:
        #     print(obj.value)

        # print(wb)

        # for rows in ws.rows:
        #     for cell in rows:
        #         print(cell.value)

        # for x in range(len(column)):
        #     print(column[x].value)

        # for rowID in range(list(ws.columns)[0].value):
        #     print(rowID)

    @staticmethod
    def activate_xlsx():
        try:
            wb = load_workbook(filename='/home/stride/PycharmProjects/TrailThings/TestDataX.xlsx')
            LOGGER.info("Workbook opened!")
            # ws = wb['TestDataS']
            return wb
        except Exception as e:
            # print("Active_xlsx:", e)
            LOGGER.debug("Invalid file")
            LOGGER.exception(e)

    @staticmethod
    def get_no_of_iteration(wb, sheet_name, col_num):
        try:
            ws = wb[sheet_name]
            row_len = len(list(ws.columns)[col_num])
            print(row_len)
            return row_len
        except Exception as e:
            print("get_no_of_iteration:", e)

    @staticmethod
    def read_excel(wb, sheet_name, row_num, col_num):
        try:
            ws = wb[sheet_name]
            cell_value = ws.cell(row=row_num, column=col_num).value
            print(cell_value)
        except Exception as e:
            print("read excel:", e)

    @staticmethod
    def write_excel(wb, sheet_name, row_num, col_num, value):
        try:
            ws = wb[sheet_name]
            ws.cell(row=row_num, column=col_num).value = value

            write_value = ws.cell(row=row_num, column=col_num).value
            write = {"value:": write_value}
            # o = write["value"]
            LOGGER.info(write['value'])
        except Exception as e:
            print("write_excel:", e)

    @staticmethod
    def save_excel(wb, file_name):
        try:
            wb.save(file_name)
        except Exception as e:
            print("save excel:", e)

    @staticmethod
    def row_list(wb, sheet_name, value):
        try:
            ws = wb[sheet_name]
            row_list = []
            for obj in list(ws.columns)[0]:
                # print(obj.row, ":", obj.value)
                if obj.value == value:
                    row_list.append(obj.row)

            print(row_list)
        except Exception as e:
            print("row_list:", e)
